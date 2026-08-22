"""Module to generate Evaluation_Metrics.xlsx report using openpyxl."""

from pathlib import Path
from typing import Any, Dict, List, Union
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def write_metrics_xlsx(
    metrics_rows: List[Dict[str, Any]],
    overall_row: Dict[str, Any],
    out_path: Union[str, Path],
) -> None:
    """Write Evaluation_Metrics.xlsx using openpyxl.

    Sheet layout:
    - Row 1 header: Sequence, Split, Sensor, TP, FP, FN, Precision, Recall, F1, AP@IoU0.5
    - One row per sequence (all 21)
    - Final row labelled OVERALL with aggregate Precision, Recall, F1 and mAP@IoU0.5
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation_Metrics"

    header_note = overall_row.get("note", "")
    if header_note:
        ws.append([header_note])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        note_cell = ws.cell(row=1, column=1)
        note_cell.font = Font(name="Calibri", size=10, italic=True, color="595959")
        note_cell.alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Sequence",
        "Split",
        "Sensor",
        "TP",
        "FP",
        "FN",
        "Precision",
        "Recall",
        "F1",
        "AP@IoU0.5",
    ]
    ws.append(headers)
    header_row_idx = ws.max_row

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row_data in metrics_rows:
        row_vals = [
            str(row_data.get("sequence", "")),
            str(row_data.get("split", "")),
            str(row_data.get("sensor", "")),
            row_data.get("tp") if row_data.get("tp") is not None else "",
            row_data.get("fp") if row_data.get("fp") is not None else "",
            row_data.get("fn") if row_data.get("fn") is not None else "",
            row_data.get("precision") if row_data.get("precision") is not None else "",
            row_data.get("recall") if row_data.get("recall") is not None else "",
            row_data.get("f1") if row_data.get("f1") is not None else "",
            row_data.get("ap") if row_data.get("ap") is not None else "",
        ]
        ws.append(row_vals)

        curr_row = ws.max_row
        for c_idx in range(1, len(headers) + 1):
            c = ws.cell(row=curr_row, column=c_idx)
            c.border = thin_border
            if c_idx == 1:
                c.alignment = left_align
            elif c_idx in (2, 3):
                c.alignment = center_align
            elif c_idx in (4, 5, 6):
                c.alignment = right_align
                if isinstance(c.value, (int, float)):
                    c.number_format = "#,##0"
            else:
                c.alignment = right_align
                if isinstance(c.value, (int, float)):
                    c.number_format = "0.000000"

    overall_vals = [
        "OVERALL",
        str(overall_row.get("split", "all")),
        "ALL",
        overall_row.get("tp") if overall_row.get("tp") is not None else "",
        overall_row.get("fp") if overall_row.get("fp") is not None else "",
        overall_row.get("fn") if overall_row.get("fn") is not None else "",
        overall_row.get("precision") if overall_row.get("precision") is not None else "",
        overall_row.get("recall") if overall_row.get("recall") is not None else "",
        overall_row.get("f1") if overall_row.get("f1") is not None else "",
        overall_row.get("mAP", overall_row.get("ap")) if overall_row.get("mAP", overall_row.get("ap")) is not None else "",
    ]
    ws.append(overall_vals)
    curr_row = ws.max_row

    overall_font = Font(name="Calibri", size=11, bold=True)
    overall_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    for c_idx in range(1, len(headers) + 1):
        c = ws.cell(row=curr_row, column=c_idx)
        c.font = overall_font
        c.fill = overall_fill
        c.border = thin_border
        if c_idx == 1:
            c.alignment = left_align
        elif c_idx in (2, 3):
            c.alignment = center_align
        elif c_idx in (4, 5, 6):
            c.alignment = right_align
            if isinstance(c.value, (int, float)):
                c.number_format = "#,##0"
        else:
            c.alignment = right_align
            if isinstance(c.value, (int, float)):
                c.number_format = "0.000000"

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(out_path)
    print(f"Metrics XLSX written successfully to: {out_path}", flush=True)
